from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color
from xml_parser import *
import os

DEFAULT_FILENAME = "test_results.xlsx"
def default_file_path():
    return os.path.join(os.path.curdir, DEFAULT_FILENAME)
SHEET_TITLE = "TEST RESULTS SHEET: "
TEST_LIST_START_INDEX = 16
FAIL_COLOUR_HEX = 'ffab97'
PASS_COLOUR_HEX = '3c5e0b'

class Xml2Excel():
    def __init__(self,  robot_results_path, xlsx_file_path=default_file_path()):
        self.results_path = robot_results_path
        self.xlxs_file_path = xlsx_file_path
        self.overall_result = True
        self.overall_result_start_row = 16
        self.overall_result_end_row = -1
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.suites = []


    def run(self):
        self.suites.append(parse_xml(self.results_path))
        # Title, Tester, Date, Approver
        self.insert_sheet_header()
        # Card/Suite Run Information
        self.insert_suite_info_headers()
        self.insert_suite_config_info()
        self.insert_overall_test_results()

        # TODO: Add Parsed Data

        self.save_workbook()

    def insert_sheet_header(self):
        sheet_title = SHEET_TITLE + self.suites[0].suite_name
        self.insert_merged_title(sheet_title, 'A1', 'E1', bold=True, font_size=14)
        self.insert_element("Staff Member:", 'A2', bold=True)
        self.insert_element(str(self.suites[0].staff_name), 'B2')
        self.insert_element("Date:", 'A3', bold=True)
        self.insert_element(date_time_formatter(self.suites[0].date_time), 'B3')
        self.insert_element("Approver:", 'A4', bold=True)
        self.insert_element("Date:", 'A5', bold=True)

    def insert_suite_config_info(self):
        for suite in self.suites:
            self.insert_element(suite.part_number, 'B8')
            self.insert_element(suite.work_order_job_number, 'B9')
            self.insert_element(suite.batch_number, 'B10')
            self.insert_element(suite.serial_number, 'B11')
            #self.insert_element(suite.suite_name, 'E7')

    def insert_overall_test_results(self):
        self.insert_overall_test_result_titles()
        pass_column = 'B'
        fail_column = 'C'
        test_name_column = 'D'
        test_list = []
        row = TEST_LIST_START_INDEX
        for test in self.suites[0].tests:
            name_cell = test_name_column + str(row)
            self.insert_element(test.name, name_cell)
            #self.insert_element(test.status, status_cell)
            if test.status == 'PASS':
                status_cell = pass_column + str(row)
                current_val = self.worksheet[status_cell]
                self.worksheet[status_cell] = current_val.value + 1
                self.colour_cell(status_cell, PASS_COLOUR_HEX)
            else:
                status_cell = fail_column + str(row)
                self.increment_cell_value(status_cell)
                self.overall_result = False
            row = row + 1

        self.overall_result_end_row = row
        self.colour_overall_result_cells()



    def insert_overall_test_result_titles(self):
        self.insert_merged_title("OVERALL RESULTS", 'A14', 'E14', bold=True, font_size=12)
        self.insert_element("Overall Status", 'A15', bold=True)
        self.insert_element("Pass", 'B15', bold=True)
        self.insert_element("Fail", 'C15', bold=True)
        self.insert_element("Test Name", 'D15', bold=True)

    def insert_suite_info_headers(self):
        self.insert_merged_title("Card Information", 'A7', 'C7', font_size=12)
        self.insert_element("Part Number:", 'A8', bold=True)
        self.insert_element("Work Order:", 'A9', bold=True)
        self.insert_element("Batch Number:", 'A10', bold=True)
        self.insert_element("Serial Number:", 'A11', bold=True)

    def colour_overall_result_cells(self):
        overall_status_column = 'A'
        pass_column = 'B'
        fail_column = 'C'

        start_row = self.overall_result_start_row
        end_row = self.overall_result_end_row

        for row in range(start_row, end_row):
            overall_status_cell = 'A' + str(row)
            pass_cell = pass_column + str(row)
            fail_cell = fail_column + str(row)
            if self.worksheet[fail_cell].value is not None and self.worksheet[fail_cell].value >= 1:
                self.colour_cell(fail_cell, FAIL_COLOUR_HEX)
                self.worksheet[overall_status_cell] = "FAIL"
                self.colour_cell(overall_status_cell, FAIL_COLOUR_HEX)
            else:
                self.colour_cell(fail_cell, FAIL_COLOUR_HEX)
                self.worksheet[overall_status_cell] = "PASS"
                self.colour_cell(overall_status_cell, PASS_COLOUR_HEX)
                self.worksheet[fail_cell] = 0
            if self.worksheet[pass_cell].value is not None and self.worksheet[pass_cell].value >= 1:
                self.colour_cell(pass_cell, PASS_COLOUR_HEX)
            else:
                self.worksheet[pass_cell] = 0




        # Output overall test result
        overall_status_row = 2 + self.overall_result_end_row
        overall_result_cell = 'B' + str(overall_status_row)
        self.insert_element("Overall Test Result:",
                            overall_status_column + str(overall_status_row), bold=True)
        if self.overall_result:
            self.insert_element("PASS", overall_result_cell)
            self.colour_cell(overall_result_cell, PASS_COLOUR_HEX)
        else:
            self.insert_element("FAIL", overall_result_cell)
            self.colour_cell(overall_result_cell, FAIL_COLOUR_HEX)

    def insert_element(self, text, element, bold=False):
        xlsx_element = self.worksheet[element]
        xlsx_element.font = Font(bold=bold)
        self.worksheet[element] = text

    def increment_cell_value(self, element, increment=1):
        if self.worksheet[element].value is None:
            self.worksheet[element] = increment
        else:
            self.worksheet[element] = self.worksheet[element].value + increment


    def insert_merged_title(self, text, start_element, end_element, bold=True, font_size=11):
        start_element_cell = self.worksheet[start_element]
        start_element_cell.font = Font(bold=bold, size=font_size)
        self.worksheet[start_element] = text
        range = str(start_element) + ':' + str(end_element)
        self.worksheet.merge_cells(range)
        self.worksheet[start_element].alignment = Alignment(horizontal='center')

        #self.worksheet.merge_cells(range)

    def colour_cell(self, element, hex_colour):
        cell = self.worksheet[element]
        colour = Color(rgb=hex_colour)
        fill = PatternFill(patternType='solid', fgColor=colour)
        cell.fill = fill

    def save_workbook(self):
        self.workbook.save(self.xlxs_file_path)

def get_xml_files(path):
    xml_files = []
    if path.endswith(".xml"):
        xml_files.append(path)
        return xml_files
    for subdir, dirs, files in os.walk(path):
        for file_name in files:
            if file_name.endswith('.xml'):
                xml_files.append(os.path.join(subdir, file_name))
    return xml_files

def date_time_formatter(date_time_str):
    year = date_time_str[0:4]
    month = date_time_str[4:6]
    day = date_time_str[6:8]
    time = date_time_str[9:-1]
    formatted_str = year + "-" + month + "-" + day + ", " + time
    return formatted_str


def input_suit_info(suite_run_ifno):
    """
    Puts suite info into excel sheet
    :param SuiteRunInfo object
    """
    pass


if __name__ == "__main__":
    #test_list = parse_xml(data_source2)
    robot_results = "H:\\builds\Capstone\\ace-test-framework\out\\bnc_card\\2021-03-08T20-16-30-048130"
    file_list = get_xml_files(data_source4)
    for file in file_list:
        print(file)
    xml_formatter = Xml2Excel(data_source3)
    xml_formatter.run()




