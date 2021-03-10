from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Color
from resultmanager.xml_parser import *
import os

DEFAULT_FILENAME = "test_results.xlsx"
def default_file_path():
    return os.path.join(os.path.curdir, DEFAULT_FILENAME)
SHEET_TITLE = "TEST RESULTS SHEET: "
TEST_LIST_START_INDEX = 16
FAIL_COLOUR_HEX = 'ffab97'
PASS_COLOUR_HEX = '3c5e0b'


class Xml2Excel:
    def __init__(self,  robot_results_path, xlsx_file_path=default_file_path()):
        self.results_path = robot_results_path
        self.xlxs_file_path = xlsx_file_path
        self.overall_result = True
        self.overall_result_start_row = 16
        self.overall_result_end_row = -1
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.suites = []


    def run(self):
        """
        Method that gets parsed xml data and inputs in in a .xlsx file

        Gets list of xml files from given file_path/dir. Uses xml_parser.py
        to get results from files. Info is formatted and output to excel file
        using openpyxl module.
        """
        # Parse xml files. Results stored in self.suites
        self.parse_results()
        # Title, Tester, Date, Approver
        self.insert_sheet_header()
        # Card/Suite Run Information
        self.insert_suite_info_headers()
        # Part number, work order, batch number, serial number, tester, date
        self.insert_suite_config_info()
        # Take results from self.suites. Input in worksheet
        self.insert_overall_test_results()
        # save into self.xlsx_file_path
        self.save_workbook()

    def insert_sheet_header(self):
        """
        Method that inputs test result info to worksheet

        Combines suite name with SHEET TITLE. Inputs parsed suite data.
        Input data includes: staff_name, date
        Template added for: Approver, Date (of approval)
        """
        sheet_title = SHEET_TITLE + self.suites[0].suite_name
        self.insert_merged_title(sheet_title, 'A1', 'E1', bold=True, font_size=14)
        self.insert_element("Staff Member:", 'A2', bold=True)
        self.insert_element(str(self.suites[0].staff_name), 'B2')
        self.insert_element("Date:", 'A3', bold=True)
        self.insert_element(date_time_formatter(self.suites[0].date_time), 'B3')
        self.insert_element("Approver:", 'A4', bold=True)
        self.insert_element("Date:", 'A5', bold=True)

    def insert_suite_config_info(self):
        """
        Method that inputs suite config info to excel worksheet

        Input data includes: part_number, work_order_job_number, batch_number,
        serial_number. Title/headings for data added in insert_suite_info_headers()
        """
        suite = self.suites[0]
        self.insert_element(suite.part_number, 'B8')
        self.insert_element(suite.work_order_job_number, 'B9')
        self.insert_element(suite.batch_number, 'B10')
        self.insert_element(suite.serial_number, 'B11')

    def insert_suite_info_headers(self):
        """
        Method that inputs suite config info into the worksheet
        """
        self.insert_merged_title("Card Information", 'A7', 'C7', font_size=12)
        self.insert_element("Part Number:", 'A8', bold=True)
        self.insert_element("Work Order:", 'A9', bold=True)
        self.insert_element("Batch Number:", 'A10', bold=True)
        self.insert_element("Serial Number:", 'A11', bold=True)

    def parse_results(self):
        """
        Method that uses xml_parser.py to get a list of
        parsed test suites

        Sets self.suites to a list of the parsed suite data
        """
        suites = []
        xml_files = get_xml_files(self.results_path)
        for file_path in xml_files:
            suites.append(parse_xml(file_path))
        self.suites = suites

    def get_test_results_dict(self):
        """
        Method that gets results from self.suites and creates a dictionary with
        test names and their corresponding pass/fail stats

        :return: Dictionary of results. Dictionary is in the form:
        key = Test Name, value = {"PASS": number of passes, "FAIL": number of fails}
        """
        test_results = {}
        suites = self.suites
        for suite in suites:
            test_list = suite.tests
            for test in test_list:
                if test.name not in test_results:
                    # insert in dictionary
                    test_results[test.name] = {"PASS": 0, "FAIL": 0}
                # increment stats depending on test status
                if test.status == 'PASS':
                    test_results[test.name]['PASS'] = test_results[test.name]['PASS'] + 1
                else:
                    self.overall_result = False
                    test_results[test.name]['FAIL'] = test_results[test.name]['FAIL'] + 1
        # TODO get suite total stats
        # TODO get suite stats by tag
        return test_results


    def insert_overall_test_results(self):
        """
        Method that inputs suite data into worksheet

        Takes dictionary from get_test_results_dict() and inputs overall test
        data from self.test_start_index till all tests and results are added.
        self.overall_result_end_row is updated to hold the bottommost row
        of data added. This is used for formatting other data
        """
        test_results = self.get_test_results_dict()
        self.insert_overall_test_result_titles()
        pass_column = 'B'
        fail_column = 'C'
        test_name_column = 'D'
        row = TEST_LIST_START_INDEX

        for test in test_results:
            pass_cell = pass_column + str(row)
            fail_cell = fail_column + str(row)
            name_cell = test_name_column + str(row)

            self.insert_element(test_results[test]['PASS'], pass_cell)
            self.insert_element(test_results[test]['FAIL'], fail_cell)
            self.insert_element(test, name_cell)
            row = row + 1

        self.overall_result_end_row = row
        self.determine_overall_results()


    def insert_overall_test_result_titles(self):
        """
        Method that inputs overall results titles info into the worksheet
        """
        self.insert_merged_title("OVERALL RESULTS", 'A14', 'E14', bold=True, font_size=12)
        self.insert_element("Overall Status", 'A15', bold=True)
        self.insert_element("Pass", 'B15', bold=True)
        self.insert_element("Fail", 'C15', bold=True)
        self.insert_element("Test Name", 'D15', bold=True)


    def determine_overall_results(self):
        """
        Method that looks through test results, determines overall result, and colours cells

        Uses self.overall_result_start_row and self.overall_result_end_row to determine
        results cell range. Colours PASS >=1 Green. Colours FAIL >= 1 red. If there are
        any fails, the overall result is FAIL
        """
        # Constants for formatting
        overall_status_column = 'A'
        pass_column = 'B'
        fail_column = 'C'
        overall_suite_status_title_column = 'A'
        overall_suite_status_column = 'B'

        # need to iterate over rows where overall results are stored
        start_row = self.overall_result_start_row
        end_row = self.overall_result_end_row

        # Iterate over result range
        for row in range(start_row, end_row):
            overall_status_cell = overall_status_column + str(row)
            pass_cell = pass_column + str(row)
            fail_cell = fail_column + str(row)
            # Check Fail Cells
            if self.worksheet[fail_cell].value >= 1:
                # Colour FAIL Cell
                self.colour_cell(fail_cell, FAIL_COLOUR_HEX)
                # Overall test status is FAIL
                self.worksheet[overall_status_cell] = "FAIL"
                self.colour_cell(overall_status_cell, FAIL_COLOUR_HEX)
            else:
                # Overall test status is pass
                self.worksheet[overall_status_cell] = "PASS"
                self.colour_cell(overall_status_cell, PASS_COLOUR_HEX)
            # Check Pass Cells
            if self.worksheet[pass_cell].value >= 1:
                self.colour_cell(pass_cell, PASS_COLOUR_HEX)

        # Output overall test result
        overall_suite_status_row = 2 + self.overall_result_end_row
        overall_result_cell = overall_suite_status_column + str(overall_suite_status_row)
        self.insert_element("Overall Test Result:",
                            overall_suite_status_title_column + str(overall_suite_status_row), bold=True)
        if self.overall_result:
            self.insert_element("PASS", overall_result_cell)
            self.colour_cell(overall_result_cell, PASS_COLOUR_HEX)
        else:
            self.insert_element("FAIL", overall_result_cell)
            self.colour_cell(overall_result_cell, FAIL_COLOUR_HEX)

    def insert_element(self, text, element, bold=False):
        """
        Helper method for inserting test into cells

        :param text: String to be inserted into cell
        :param element: String representing cell element. E.g. 'A1'
        :param bold: If true determines if the font is bolded
        """
        xlsx_element = self.worksheet[element]
        xlsx_element.font = Font(bold=bold)
        self.worksheet[element] = text

    def increment_cell_value(self, element, increment=1):
        """
        Helper method for incrementing value in cell

        :param element: String representing cell element. E.g. 'A1'
        :param increment: number cell value will be increased by increment value
        """
        if self.worksheet[element].value is None:
            self.worksheet[element] = increment
        else:
            self.worksheet[element] = self.worksheet[element].value + increment


    def insert_merged_title(self, text, start_element, end_element, bold=True, font_size=11):
        start_element_cell = self.worksheet[start_element]
        start_element_cell.font = Font(bold=bold, size=font_size)
        self.worksheet[start_element] = text
        range = str(start_element) + ':' + str(end_element)
        self.worksheet.merge_cells(range)
        self.worksheet[start_element].alignment = Alignment(horizontal='center')

        #self.worksheet.merge_cells(range)

    def colour_cell(self, element, hex_colour):
        """
        Helper Function for colouring cells

        :param element: String representing cell element. E.g. 'A1'
        :param hex_colour: String representing hex colour. E.g '3c5e0b'
        """
        cell = self.worksheet[element]
        colour = Color(rgb=hex_colour)
        fill = PatternFill(patternType='solid', fgColor=colour)
        cell.fill = fill

    def save_workbook(self):
        """
        Method that saves self.workbook using self.xlsx_file_path
        """
        self.workbook.save(self.xlxs_file_path)

def get_xml_files(path):
    """
    Function for getting all .xml files in a given directory
    If .xml file path is passed instead of a directory, method will just return
    .xml file path

    :param path: String pointing to device directory or .xml file
    :return: list of .xml file paths
    """
    xml_files = []
    if path.endswith(".xml"):
        xml_files.append(path)
        return xml_files
    for subdir, dirs, files in os.walk(path):
        for file_name in files:
            if file_name.endswith('.xml'):
                xml_files.append(os.path.join(subdir, file_name))
    return xml_files

def date_time_formatter(date_time_str):
    """
    Helper function for formatting data from parsed xml file

    :param date_time_str: String
    :return:
    """
    year = date_time_str[0:4]
    month = date_time_str[4:6]
    day = date_time_str[6:8]
    time = date_time_str[9:-1]
    formatted_str = year + "-" + month + "-" + day + ", " + time
    return formatted_str


if __name__ == "__main__":
    xml_formatter = Xml2Excel(data_source3)
    xml_formatter.run()




